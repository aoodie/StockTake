from __future__ import annotations

from io import BytesIO
from sqlite3 import Connection

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


EXPORT_COLUMNS = [
    "Session ID",
    "Session Name",
    "Location",
    "BIN",
    "Barcode",
    "Product Name",
    "Category",
    "Size",
    "Quantity",
    "Unit",
    "Draft Status",
    "Missing BIN Flag",
    "Counted At",
    "Device ID",
    "Notes",
    "Case Type",
]

def spreadsheet_safe(value):
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


def build_stocktake_workbook(db: Connection, session_id: str, *, prefer_scan_snapshots: bool = False) -> bytes:
    workbook = Workbook()
    stocktake = workbook.active
    stocktake.title = "Stocktake"
    exceptions = workbook.create_sheet("Missing BIN Exceptions")

    for sheet in (stocktake, exceptions):
        sheet.append(EXPORT_COLUMNS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF7")

    bin_value = "COALESCE(sl.bin_snapshot, p.bin, '')" if prefer_scan_snapshots else "COALESCE(p.bin, sl.bin_snapshot, '')"
    barcode_value = "COALESCE(sl.barcode_snapshot, p.barcode, '')" if prefer_scan_snapshots else "COALESCE(p.barcode, sl.barcode_snapshot, '')"
    name_value = "COALESCE(sl.product_name_snapshot, p.name, '')" if prefer_scan_snapshots else "COALESCE(p.name, sl.product_name_snapshot, '')"
    draft_value = "COALESCE(sl.draft_status, p.draft_status, 'confirmed')" if prefer_scan_snapshots else "COALESCE(p.draft_status, sl.draft_status, 'confirmed')"
    rows = db.execute(
        f"""
        SELECT
            sl.session_id,
            COALESCE(s.name, sl.session_id) AS session_name,
            COALESCE(l.name, sl.location_id) AS location_name,
            {bin_value} AS bin,
            {barcode_value} AS barcode,
            {name_value} AS product_name,
            COALESCE(p.category, '') AS category,
            COALESCE(p.size, '') AS size,
            sl.quantity_decimal,
            COALESCE(p.unit, 'each') AS unit,
            {draft_value} AS draft_status,
            CASE WHEN {bin_value} = '' THEN 'Y' ELSE 'N' END AS missing_bin,
            sl.counted_at,
            sl.device_id,
            COALESCE(sl.notes, '') AS notes,
            COALESCE(sl.case_type, 'split') AS case_type
        FROM stocktake_lines sl
        LEFT JOIN products p ON p.id = sl.product_id
        LEFT JOIN sessions s ON s.id = sl.session_id
        LEFT JOIN locations l ON l.id = sl.location_id
        WHERE sl.session_id = ?
        ORDER BY l.name, p.name, sl.counted_at
        """,
        (session_id,),
    ).fetchall()

    for row in rows:
        values = [spreadsheet_safe(value) for value in row]
        # Preserve explicit zero as "0"; never convert falsey quantities to blanks.
        if values[8] == "":
            values[8] = "0"
        stocktake.append(values)
        if values[11] == "Y":
            exceptions.append(values)

    for sheet in (stocktake, exceptions):
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 40)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
