import sqlite3
from io import BytesIO

from openpyxl import load_workbook

from app.exporter import EXPORT_COLUMNS, build_stocktake_workbook


def test_export_preserves_zero_and_missing_bin_exception():
    db = sqlite3.connect(":memory:")
    db.executescript(
        """
        CREATE TABLE products (
            id TEXT PRIMARY KEY,
            barcode TEXT,
            bin TEXT,
            name TEXT,
            category TEXT,
            size TEXT,
            unit TEXT,
            draft_status TEXT
        );
        CREATE TABLE sessions (id TEXT PRIMARY KEY, name TEXT);
        CREATE TABLE locations (id TEXT PRIMARY KEY, name TEXT);
        CREATE TABLE stocktake_lines (
            id TEXT PRIMARY KEY,
            session_id TEXT,
            location_id TEXT,
            product_id TEXT,
            barcode_snapshot TEXT,
            bin_snapshot TEXT,
            product_name_snapshot TEXT,
            quantity_decimal TEXT,
            case_type TEXT,
            draft_status TEXT,
            counted_at TEXT,
            device_id TEXT,
            notes TEXT
        );
        """
    )
    db.execute("INSERT INTO sessions VALUES ('s1', 'May 27')")
    db.execute("INSERT INTO locations VALUES ('l1', 'Main Bar')")
    db.execute(
        "INSERT INTO products VALUES ('p1', '123', '', 'Gin', 'Spirit', '70cl', 'bottle', 'confirmed')"
    )
    db.execute(
        """
        INSERT INTO stocktake_lines VALUES (
            'line1', 's1', 'l1', 'p1', '123', '', 'Gin', '0', 'split', 'confirmed',
            '2026-05-27T10:00:00Z', 'device-a', ''
        )
        """
    )

    data = build_stocktake_workbook(db, "s1")
    workbook = load_workbook(BytesIO(data))
    assert workbook["Stocktake"][1][0].value == EXPORT_COLUMNS[0]
    assert workbook["Stocktake"][2][8].value == "0"
    assert workbook["Stocktake"][2][9].value == "split unit"
    assert workbook["Stocktake"][2][15].value == "split"
    assert workbook["Missing BIN Exceptions"][2][8].value == "0"
    assert workbook["Missing BIN Exceptions"][2][11].value == "Y"


def test_raw_export_prefers_barcode_and_name_captured_at_scan_time():
    db = sqlite3.connect(":memory:")
    db.executescript(
        """
        CREATE TABLE products (
            id TEXT PRIMARY KEY, barcode TEXT, bin TEXT, name TEXT, category TEXT,
            size TEXT, unit TEXT, draft_status TEXT
        );
        CREATE TABLE sessions (id TEXT PRIMARY KEY, name TEXT);
        CREATE TABLE locations (id TEXT PRIMARY KEY, name TEXT);
        CREATE TABLE stocktake_lines (
            id TEXT PRIMARY KEY, session_id TEXT, location_id TEXT, product_id TEXT,
            barcode_snapshot TEXT, bin_snapshot TEXT, product_name_snapshot TEXT,
            quantity_decimal TEXT, case_type TEXT, draft_status TEXT, counted_at TEXT, device_id TEXT, notes TEXT
        );
        """
    )
    db.execute("INSERT INTO products VALUES ('pw-1', 'PW-PID', 'PW-BIN', 'Mapped PW Name', '', '', 'case', 'confirmed')")
    db.execute(
        """
        INSERT INTO stocktake_lines VALUES (
            'line1', 's1', 'l1', 'pw-1', 'PHYSICAL-BARCODE', '', 'Name Seen During Scan', '2', 'full', 'draft',
            '2026-06-07T10:00:00Z', 'phone-a', ''
        )
        """
    )

    data = build_stocktake_workbook(db, "s1", prefer_scan_snapshots=True)
    workbook = load_workbook(BytesIO(data))
    row = workbook["Stocktake"][2]

    assert row[4].value == "PHYSICAL-BARCODE"
    assert row[5].value == "Name Seen During Scan"
    assert row[9].value == "full case"
    assert row[10].value == "draft"
    assert row[15].value == "full"
